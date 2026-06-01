"""Dispatcher-style local review table contracts and helpers.

Dispatcher review rows may contain private predicted values when explicitly
written to ignored local files. Feedback rows and aggregates are safe summaries:
they keep only booleans, counts, aliases, field names, and issue types.
"""

from collections import Counter
import csv
from pathlib import Path

from app.document_ai.private_measurement_outputs import (
    DEFAULT_PRIVATE_MEASUREMENT_OUTPUT_DIR,
    _normalize_output_dir,
)

from app.document_ai.review_issue_taxonomy import (
    REVIEW_ISSUE_TYPE_BROKER_MISSING,
    REVIEW_ISSUE_TYPE_EXTRA_VALUE,
    REVIEW_ISSUE_TYPE_LOAD_ID_MISSING,
    REVIEW_ISSUE_TYPE_MISSING_VALUE,
    REVIEW_ISSUE_TYPE_OTHER,
    REVIEW_ISSUE_TYPE_WRONG_BROKER,
    REVIEW_ISSUE_TYPE_WRONG_DATE,
    REVIEW_ISSUE_TYPE_WRONG_DELIVERY,
    REVIEW_ISSUE_TYPE_WRONG_LOAD_ID,
    REVIEW_ISSUE_TYPE_WRONG_PICKUP,
    REVIEW_ISSUE_TYPE_WRONG_RATE,
    REVIEW_ISSUE_TYPE_WRONG_VALUE,
    normalize_review_issue_type,
)
from app.document_ai.review_feedback_target_selector import (
    select_repair_target_from_dispatcher_feedback,
)


DISPATCHER_REVIEW_TABLE_VERSION = "dispatcher_review_table_v3"

DISPATCHER_REVIEW_V3_WORKBOOK_XLSX = "ratecon_review_v3_dispatcher_workbook.xlsx"
DISPATCHER_REVIEW_V3_REVIEW_CSV = "ratecon_review_v3_dispatcher_review.csv"
DISPATCHER_REVIEW_V3_AUDIT_CSV = "ratecon_review_v3_extraction_audit.csv"
DISPATCHER_REVIEW_V3_INSTRUCTIONS_CSV = "ratecon_review_v3_instructions.csv"
DISPATCHER_REVIEW_V3_FEEDBACK_SUMMARY_CSV = (
    "ratecon_review_v3_feedback_summary.csv"
)

SHEET_DISPATCHER_REVIEW = "Dispatcher_Review"
SHEET_EXTRACTION_AUDIT = "Extraction_Audit"
SHEET_REVIEW_INSTRUCTIONS = "Review_Instructions"
SHEET_FEEDBACK_SUMMARY = "Feedback_Summary"

DISPATCHER_FIELD_BROKER = "broker"
DISPATCHER_FIELD_PICKUP = "pickup"
DISPATCHER_FIELD_PICKUP_DATE = "pickup_date"
DISPATCHER_FIELD_DELIVERY = "delivery"
DISPATCHER_FIELD_DELIVERY_DATE = "delivery_date"
DISPATCHER_FIELD_LOAD_NUMBER = "load_number"
DISPATCHER_FIELD_CARRIER = "carrier"
DISPATCHER_FIELD_TRAILER_TYPE = "trailer_type"
DISPATCHER_FIELD_COMMODITY = "commodity"
DISPATCHER_FIELD_TOTAL_WEIGHT = "total_weight"
DISPATCHER_FIELD_FINAL_RATE = "final_rate"
DISPATCHER_FIELD_SPECIAL_REQUIREMENTS = "special_requirements"

DISPATCHER_EDITABLE_FIELDS = [
    DISPATCHER_FIELD_BROKER,
    DISPATCHER_FIELD_PICKUP,
    DISPATCHER_FIELD_PICKUP_DATE,
    DISPATCHER_FIELD_DELIVERY,
    DISPATCHER_FIELD_DELIVERY_DATE,
    DISPATCHER_FIELD_LOAD_NUMBER,
    DISPATCHER_FIELD_CARRIER,
    DISPATCHER_FIELD_TRAILER_TYPE,
    DISPATCHER_FIELD_COMMODITY,
    DISPATCHER_FIELD_TOTAL_WEIGHT,
    DISPATCHER_FIELD_FINAL_RATE,
    DISPATCHER_FIELD_SPECIAL_REQUIREMENTS,
]

DISPATCHER_REVIEW_COLUMNS = [
    "Folder Order",
    "Local Document Name / File Stem",
    "Measurement Alias",
    "Document Type",
    "OCR Needed",
    "Extraction Relevant",
    "Readiness Level",
    "Review Priority",
    "Source",
    "Broker",
    "Pickup",
    "Pickup Date",
    "Delivery",
    "Delivery Date",
    "Load No",
    "Carrier",
    "Trailer Type",
    "Commodity",
    "Total Weight",
    "Final Rate",
    "Special Requirements",
    "Top Blockers",
    "Predicted Status Summary",
    "User Review Status",
    "User Notes Local Only",
    "User Corrected Broker",
    "User Corrected Pickup",
    "User Corrected Pickup Date",
    "User Corrected Delivery",
    "User Corrected Delivery Date",
    "User Corrected Load No",
    "User Corrected Final Rate",
]

DISPATCHER_AUDIT_COLUMNS = [
    "Measurement Alias",
    "Field Name",
    "Predicted Value LOCAL ONLY",
    "Dispatcher Value At Export LOCAL ONLY",
    "Predicted Status",
    "Candidate Count",
    "Conflict Reason",
    "Gap Reason",
    "Source Sheet",
    "Source Field",
    "Blocker Type",
    "Warning Codes",
]

DISPATCHER_INSTRUCTIONS_COLUMNS = ["Section", "Instruction"]

DISPATCHER_FEEDBACK_SUMMARY_COLUMNS = ["Metric", "Value"]

DISPATCHER_REVIEW_COLUMNS_BY_SHEET = {
    SHEET_DISPATCHER_REVIEW: DISPATCHER_REVIEW_COLUMNS,
    SHEET_EXTRACTION_AUDIT: DISPATCHER_AUDIT_COLUMNS,
    SHEET_REVIEW_INSTRUCTIONS: DISPATCHER_INSTRUCTIONS_COLUMNS,
    SHEET_FEEDBACK_SUMMARY: DISPATCHER_FEEDBACK_SUMMARY_COLUMNS,
}

FIELD_TO_REVIEW_COLUMN = {
    DISPATCHER_FIELD_BROKER: "Broker",
    DISPATCHER_FIELD_PICKUP: "Pickup",
    DISPATCHER_FIELD_PICKUP_DATE: "Pickup Date",
    DISPATCHER_FIELD_DELIVERY: "Delivery",
    DISPATCHER_FIELD_DELIVERY_DATE: "Delivery Date",
    DISPATCHER_FIELD_LOAD_NUMBER: "Load No",
    DISPATCHER_FIELD_CARRIER: "Carrier",
    DISPATCHER_FIELD_TRAILER_TYPE: "Trailer Type",
    DISPATCHER_FIELD_COMMODITY: "Commodity",
    DISPATCHER_FIELD_TOTAL_WEIGHT: "Total Weight",
    DISPATCHER_FIELD_FINAL_RATE: "Final Rate",
    DISPATCHER_FIELD_SPECIAL_REQUIREMENTS: "Special Requirements",
}

FIELD_TO_CORRECTION_COLUMN = {
    field_name: f"User Corrected {column_name}"
    for field_name, column_name in FIELD_TO_REVIEW_COLUMN.items()
}

CORE_FIELD_TO_DISPATCHER_FIELD = {
    "broker_name": DISPATCHER_FIELD_BROKER,
    "pickup_location": DISPATCHER_FIELD_PICKUP,
    "pickup_date": DISPATCHER_FIELD_PICKUP_DATE,
    "delivery_location": DISPATCHER_FIELD_DELIVERY,
    "delivery_date": DISPATCHER_FIELD_DELIVERY_DATE,
    "load_number": DISPATCHER_FIELD_LOAD_NUMBER,
    "rate": DISPATCHER_FIELD_FINAL_RATE,
}

DETAIL_FIELD_TO_DISPATCHER_FIELD = {
    "carrier": DISPATCHER_FIELD_CARRIER,
    "carrier_name": DISPATCHER_FIELD_CARRIER,
    "equipment": DISPATCHER_FIELD_TRAILER_TYPE,
    "trailer_type": DISPATCHER_FIELD_TRAILER_TYPE,
    "commodity": DISPATCHER_FIELD_COMMODITY,
    "weight": DISPATCHER_FIELD_TOTAL_WEIGHT,
    "total_weight": DISPATCHER_FIELD_TOTAL_WEIGHT,
    "special_requirements": DISPATCHER_FIELD_SPECIAL_REQUIREMENTS,
    "requirements": DISPATCHER_FIELD_SPECIAL_REQUIREMENTS,
    "notes": DISPATCHER_FIELD_SPECIAL_REQUIREMENTS,
}

FIELD_TO_DEFAULT_ISSUE = {
    DISPATCHER_FIELD_BROKER: REVIEW_ISSUE_TYPE_WRONG_BROKER,
    DISPATCHER_FIELD_PICKUP: REVIEW_ISSUE_TYPE_WRONG_PICKUP,
    DISPATCHER_FIELD_PICKUP_DATE: REVIEW_ISSUE_TYPE_WRONG_DATE,
    DISPATCHER_FIELD_DELIVERY: REVIEW_ISSUE_TYPE_WRONG_DELIVERY,
    DISPATCHER_FIELD_DELIVERY_DATE: REVIEW_ISSUE_TYPE_WRONG_DATE,
    DISPATCHER_FIELD_LOAD_NUMBER: REVIEW_ISSUE_TYPE_WRONG_LOAD_ID,
    DISPATCHER_FIELD_FINAL_RATE: REVIEW_ISSUE_TYPE_WRONG_RATE,
}

FIELD_TO_MISSING_ISSUE = {
    DISPATCHER_FIELD_BROKER: REVIEW_ISSUE_TYPE_BROKER_MISSING,
    DISPATCHER_FIELD_LOAD_NUMBER: REVIEW_ISSUE_TYPE_LOAD_ID_MISSING,
    DISPATCHER_FIELD_FINAL_RATE: REVIEW_ISSUE_TYPE_WRONG_RATE,
}


def _text(value):
    return str(value or "").strip()


def _token(value):
    return _text(value).lower().replace(" ", "_").replace("-", "_")


def _bool_text(value):
    return "yes" if bool(value) else "no"


def _yes(value):
    return _token(value) in {"yes", "true", "1"}


def _normalize_compare(value):
    return " ".join(_text(value).split()).casefold()


def _private_value(row, include_private_values=False):
    if not include_private_values:
        return ""
    return _text((row or {}).get("Predicted Value LOCAL ONLY"))


def _status_needs_blank(status):
    return _token(status) in {"missing", "conflict", "low_confidence"}


def _candidate_count(row):
    for key in (
        "Candidate Count",
        "Load Identifier Candidate Count",
        "Main Rate Candidate Count",
    ):
        value = _text((row or {}).get(key))
        if value:
            return value
    return ""


def _gap_reason(row):
    for key in ("Gap Reason", "Policy Gap Reason", "Load Identifier Gap Reason"):
        value = _text((row or {}).get(key))
        if value:
            return value
    return ""


def _top_blockers(row):
    for key in ("Top Blockers", "Top Blocker"):
        value = _text((row or {}).get(key))
        if value:
            return value
    return ""


def _prediction_status_summary(predictions):
    parts = []
    for field_name in DISPATCHER_EDITABLE_FIELDS:
        status = _text((predictions.get(field_name) or {}).get("status"))
        if status:
            parts.append(f"{field_name}={status}")
    return ";".join(parts)


def build_dispatcher_review_row(
    folder_order="",
    local_document_name_or_file_stem="",
    measurement_alias="",
    document_type="",
    ocr_needed=False,
    extraction_relevant=True,
    readiness_level="",
    review_priority="",
    source="local_review_outputs",
    broker="",
    pickup="",
    pickup_date="",
    delivery="",
    delivery_date="",
    load_number="",
    carrier="",
    trailer_type="",
    commodity="",
    total_weight="",
    final_rate="",
    special_requirements="",
    top_blockers="",
    predicted_status_summary="",
    user_review_status="",
    user_notes_local_only="",
):
    return {
        "Folder Order": folder_order,
        "Local Document Name / File Stem": _text(local_document_name_or_file_stem),
        "Measurement Alias": _text(measurement_alias),
        "Document Type": _text(document_type),
        "OCR Needed": _bool_text(ocr_needed),
        "Extraction Relevant": _bool_text(extraction_relevant),
        "Readiness Level": _text(readiness_level),
        "Review Priority": _text(review_priority),
        "Source": _text(source),
        "Broker": _text(broker),
        "Pickup": _text(pickup),
        "Pickup Date": _text(pickup_date),
        "Delivery": _text(delivery),
        "Delivery Date": _text(delivery_date),
        "Load No": _text(load_number),
        "Carrier": _text(carrier),
        "Trailer Type": _text(trailer_type),
        "Commodity": _text(commodity),
        "Total Weight": _text(total_weight),
        "Final Rate": _text(final_rate),
        "Special Requirements": _text(special_requirements),
        "Top Blockers": _text(top_blockers),
        "Predicted Status Summary": _text(predicted_status_summary),
        "User Review Status": _text(user_review_status),
        "User Notes Local Only": _text(user_notes_local_only),
        "User Corrected Broker": "",
        "User Corrected Pickup": "",
        "User Corrected Pickup Date": "",
        "User Corrected Delivery": "",
        "User Corrected Delivery Date": "",
        "User Corrected Load No": "",
        "User Corrected Final Rate": "",
    }


def build_dispatcher_audit_row(
    measurement_alias="",
    field_name="",
    predicted_value_local_only="",
    dispatcher_value_at_export_local_only="",
    predicted_status="",
    candidate_count="",
    conflict_reason="",
    gap_reason="",
    source_sheet="",
    source_field="",
    blocker_type="",
    warning_codes=None,
):
    return {
        "Measurement Alias": _text(measurement_alias),
        "Field Name": _token(field_name),
        "Predicted Value LOCAL ONLY": _text(predicted_value_local_only),
        "Dispatcher Value At Export LOCAL ONLY": _text(
            dispatcher_value_at_export_local_only
        ),
        "Predicted Status": _text(predicted_status),
        "Candidate Count": _text(candidate_count),
        "Conflict Reason": _text(conflict_reason),
        "Gap Reason": _text(gap_reason),
        "Source Sheet": _text(source_sheet),
        "Source Field": _text(source_field),
        "Blocker Type": _text(blocker_type),
        "Warning Codes": ";".join(_token(code) for code in warning_codes or [] if _token(code)),
    }


def _prediction_from_row(
    row,
    field_name,
    include_private_values=False,
    source_sheet="",
    source_field="",
):
    status = _text((row or {}).get("Predicted Status") or (row or {}).get("Status"))
    return {
        "field_name": field_name,
        "value": _private_value(row, include_private_values=include_private_values),
        "status": status,
        "candidate_count": _candidate_count(row),
        "conflict_reason": _text(
            (row or {}).get("Conflict Reason") or (row or {}).get("Rate Conflict Reason")
        ),
        "gap_reason": _gap_reason(row),
        "source_sheet": source_sheet,
        "source_field": source_field or _text((row or {}).get("Field Name")),
        "blocker_type": "review_needed"
        if _token(status) in {"missing", "conflict", "needs_review", "low_confidence"}
        else "",
    }


def _set_prediction(predictions, dispatcher_field, prediction):
    current = predictions.get(dispatcher_field)
    if not current:
        predictions[dispatcher_field] = prediction
        return
    current_status = _token(current.get("status"))
    new_status = _token(prediction.get("status"))
    if current_status in {"missing", ""} and new_status not in {"missing", ""}:
        predictions[dispatcher_field] = prediction


def _field_rows_by_alias(rows):
    by_alias = {}
    for row in rows or []:
        alias = _text((row or {}).get("Measurement Alias"))
        if alias:
            by_alias.setdefault(alias, []).append(row)
    return by_alias


def _document_sort_key(row):
    try:
        order = int(_text((row or {}).get("Folder Order")) or 0)
    except ValueError:
        order = 0
    return (order, _text((row or {}).get("Measurement Alias")))


def build_dispatcher_review_table_from_rows(
    document_rows,
    core_field_rows=None,
    stop_rows=None,
    rate_rows=None,
    load_id_rows=None,
    detailed_field_rows=None,
    include_private_values=False,
):
    core_by_alias = _field_rows_by_alias(core_field_rows)
    stop_by_alias = _field_rows_by_alias(stop_rows)
    rate_by_alias = _field_rows_by_alias(rate_rows)
    load_id_by_alias = _field_rows_by_alias(load_id_rows)
    detail_by_alias = _field_rows_by_alias(detailed_field_rows)
    dispatcher_rows = []
    audit_rows = []

    for doc in sorted(document_rows or [], key=_document_sort_key):
        alias = _text(doc.get("Measurement Alias"))
        predictions = {}

        for row in core_by_alias.get(alias, []):
            field_name = _token(row.get("Field Name"))
            dispatcher_field = CORE_FIELD_TO_DISPATCHER_FIELD.get(field_name)
            if not dispatcher_field:
                continue
            _set_prediction(
                predictions,
                dispatcher_field,
                _prediction_from_row(
                    row,
                    dispatcher_field,
                    include_private_values=include_private_values,
                    source_sheet="Core_Field_Review",
                    source_field=field_name,
                ),
            )

        for row in detail_by_alias.get(alias, []):
            field_name = _token(row.get("Field Name"))
            dispatcher_field = DETAIL_FIELD_TO_DISPATCHER_FIELD.get(field_name)
            if not dispatcher_field:
                continue
            _set_prediction(
                predictions,
                dispatcher_field,
                _prediction_from_row(
                    row,
                    dispatcher_field,
                    include_private_values=include_private_values,
                    source_sheet="Field_Review",
                    source_field=field_name,
                ),
            )

        for row in stop_by_alias.get(alias, []):
            stop_type = _token(row.get("Stop Type"))
            field_name = _token(row.get("Field Name"))
            dispatcher_field = ""
            if stop_type == "pickup" and field_name in {"location", "pickup_location"}:
                dispatcher_field = DISPATCHER_FIELD_PICKUP
            elif stop_type == "pickup" and field_name in {"date", "pickup_date"}:
                dispatcher_field = DISPATCHER_FIELD_PICKUP_DATE
            elif stop_type == "delivery" and field_name in {"location", "delivery_location"}:
                dispatcher_field = DISPATCHER_FIELD_DELIVERY
            elif stop_type == "delivery" and field_name in {"date", "delivery_date"}:
                dispatcher_field = DISPATCHER_FIELD_DELIVERY_DATE
            if dispatcher_field:
                _set_prediction(
                    predictions,
                    dispatcher_field,
                    _prediction_from_row(
                        row,
                        dispatcher_field,
                        include_private_values=include_private_values,
                        source_sheet="Stop_Review",
                        source_field=field_name,
                    ),
                )

        for row in rate_by_alias.get(alias, []):
            _set_prediction(
                predictions,
                DISPATCHER_FIELD_FINAL_RATE,
                _prediction_from_row(
                    row,
                    DISPATCHER_FIELD_FINAL_RATE,
                    include_private_values=include_private_values,
                    source_sheet="Rate_Review",
                    source_field=_text(row.get("Rate Candidate Type")) or "rate",
                ),
            )

        for row in load_id_by_alias.get(alias, []):
            if _yes(row.get("Primary Candidate?")) or _token(row.get("Identifier Type")) == "load_number":
                _set_prediction(
                    predictions,
                    DISPATCHER_FIELD_LOAD_NUMBER,
                    _prediction_from_row(
                        row,
                        DISPATCHER_FIELD_LOAD_NUMBER,
                        include_private_values=include_private_values,
                        source_sheet="Load_ID_Review",
                        source_field=_text(row.get("Identifier Type")) or "load_number",
                    ),
                )

        def value_for(field_name):
            prediction = predictions.get(field_name) or {}
            if _status_needs_blank(prediction.get("status")):
                return ""
            return prediction.get("value", "")

        dispatcher_rows.append(
            build_dispatcher_review_row(
                folder_order=_text(doc.get("Folder Order")),
                local_document_name_or_file_stem=_text(
                    doc.get("Local Document Name / File Stem")
                ),
                measurement_alias=alias,
                document_type=_text(doc.get("Document Type")),
                ocr_needed=_yes(doc.get("OCR Needed")),
                extraction_relevant=_yes(doc.get("Extraction Relevant")),
                readiness_level=_text(doc.get("Readiness Level")),
                review_priority=_text(doc.get("Review Priority")),
                source="local_review_outputs",
                broker=value_for(DISPATCHER_FIELD_BROKER),
                pickup=value_for(DISPATCHER_FIELD_PICKUP),
                pickup_date=value_for(DISPATCHER_FIELD_PICKUP_DATE),
                delivery=value_for(DISPATCHER_FIELD_DELIVERY),
                delivery_date=value_for(DISPATCHER_FIELD_DELIVERY_DATE),
                load_number=value_for(DISPATCHER_FIELD_LOAD_NUMBER),
                carrier=value_for(DISPATCHER_FIELD_CARRIER),
                trailer_type=value_for(DISPATCHER_FIELD_TRAILER_TYPE),
                commodity=value_for(DISPATCHER_FIELD_COMMODITY),
                total_weight=value_for(DISPATCHER_FIELD_TOTAL_WEIGHT),
                final_rate=value_for(DISPATCHER_FIELD_FINAL_RATE),
                special_requirements=value_for(DISPATCHER_FIELD_SPECIAL_REQUIREMENTS),
                top_blockers=_top_blockers(doc),
                predicted_status_summary=_prediction_status_summary(predictions),
                user_review_status="",
                user_notes_local_only="",
            )
        )

        for field_name in DISPATCHER_EDITABLE_FIELDS:
            prediction = predictions.get(field_name) or {"field_name": field_name}
            audit_rows.append(
                build_dispatcher_audit_row(
                    measurement_alias=alias,
                    field_name=field_name,
                    predicted_value_local_only=prediction.get("value", ""),
                    dispatcher_value_at_export_local_only=value_for(field_name),
                    predicted_status=prediction.get("status", ""),
                    candidate_count=prediction.get("candidate_count", ""),
                    conflict_reason=prediction.get("conflict_reason", ""),
                    gap_reason=prediction.get("gap_reason", ""),
                    source_sheet=prediction.get("source_sheet", ""),
                    source_field=prediction.get("source_field", ""),
                    blocker_type=prediction.get("blocker_type", ""),
                )
            )

    return {
        "dispatcher_rows": dispatcher_rows,
        "audit_rows": audit_rows,
        "summary": {
            "document_rows": len(dispatcher_rows),
            "audit_rows": len(audit_rows),
            "include_private_values_local_only": bool(include_private_values),
            "private_values_printed": False,
            "raw_text_printed": False,
            "money_values_printed": False,
        },
    }


def infer_dispatcher_issue_type(field_name, original_value, corrected_value):
    field = _token(field_name)
    original_present = bool(_text(original_value))
    corrected_present = bool(_text(corrected_value))
    if not original_present and corrected_present:
        return FIELD_TO_MISSING_ISSUE.get(field, REVIEW_ISSUE_TYPE_MISSING_VALUE)
    if original_present and not corrected_present:
        return REVIEW_ISSUE_TYPE_EXTRA_VALUE
    if original_present and corrected_present:
        return FIELD_TO_DEFAULT_ISSUE.get(field, REVIEW_ISSUE_TYPE_WRONG_VALUE)
    return REVIEW_ISSUE_TYPE_OTHER


def build_dispatcher_feedback_row(
    measurement_alias="",
    field_name="",
    original_predicted_value="",
    user_corrected_value="",
    user_issue_type="",
    user_review_status="",
    user_notes_local_only="",
    warning_codes=None,
):
    original_present = bool(_text(original_predicted_value))
    corrected_present = bool(_text(user_corrected_value))
    changed = _normalize_compare(original_predicted_value) != _normalize_compare(
        user_corrected_value
    )
    issue_type = normalize_review_issue_type(user_issue_type)
    if changed and not issue_type:
        issue_type = infer_dispatcher_issue_type(
            field_name,
            original_predicted_value,
            user_corrected_value,
        )
    return {
        "measurement_alias": _text(measurement_alias),
        "field_name": _token(field_name),
        "original_predicted_present": original_present,
        "user_corrected_present": corrected_present,
        "changed": changed,
        "inferred_issue_type": issue_type,
        "user_review_status": _token(user_review_status),
        "user_notes_present": bool(_text(user_notes_local_only)),
        "warning_codes": [_token(code) for code in warning_codes or [] if _token(code)],
        "private_values_included": False,
        "raw_text_included": False,
        "money_values_included": False,
    }


def _top_counts(counts):
    return dict(sorted(counts.items(), key=lambda item: (-item[1], item[0])))


def aggregate_dispatcher_feedback(feedback_rows):
    safe_rows = [row for row in feedback_rows or [] if isinstance(row, dict)]
    changed_rows = [row for row in safe_rows if row.get("changed")]
    issue_counts = Counter(
        row.get("inferred_issue_type") or REVIEW_ISSUE_TYPE_OTHER
        for row in changed_rows
    )
    field_counts = Counter(row.get("field_name") or "unknown" for row in changed_rows)
    changed_aliases = sorted(
        {
            row.get("measurement_alias")
            for row in changed_rows
            if row.get("measurement_alias")
        }
    )
    feedback_like_aggregate = {
        "rows_loaded": len(safe_rows),
        "changed_field_count": len(changed_rows),
        "issue_type_counts": dict(issue_counts),
    }
    decision = select_repair_target_from_dispatcher_feedback(feedback_like_aggregate)
    return {
        "rows_loaded": len(safe_rows),
        "documents_reviewed": len(
            {
                row.get("measurement_alias")
                for row in safe_rows
                if row.get("measurement_alias")
            }
        ),
        "changed_field_count": len(changed_rows),
        "issue_type_counts": _top_counts(issue_counts),
        "changed_fields_by_name": _top_counts(field_counts),
        "changed_aliases": changed_aliases,
        "recommended_next_repair_target": decision.get("selected_target"),
        "analysis_version": DISPATCHER_REVIEW_TABLE_VERSION,
        "private_values_included": False,
        "raw_text_included": False,
        "money_values_included": False,
    }


def dispatcher_instruction_rows(include_private_values=False):
    return [
        {
            "Section": "Review",
            "Instruction": "Edit Dispatcher_Review values directly or use User Corrected columns for corrections.",
        },
        {
            "Section": "Audit",
            "Instruction": "Extraction_Audit preserves original predictions and statuses for feedback import. Do not edit it.",
        },
        {
            "Section": "Private values",
            "Instruction": (
                "Predicted values are included for local review only."
                if include_private_values
                else "Predicted values are blank in status-only mode."
            ),
        },
        {
            "Section": "Do not share",
            "Instruction": "Do not share predicted values, corrected values, raw text, private filenames, local paths, money amounts, or service account keys.",
        },
        {
            "Section": "No automation",
            "Instruction": "This table does not create DispatchCases, decisions, Telegram messages, or event timeline entries.",
        },
    ]


def dispatcher_feedback_summary_rows(feedback_aggregate=None):
    aggregate = feedback_aggregate or {}
    rows = [
        {"Metric": "rows_loaded", "Value": aggregate.get("rows_loaded", 0)},
        {
            "Metric": "changed_field_count",
            "Value": aggregate.get("changed_field_count", 0),
        },
        {
            "Metric": "recommended_next_repair_target",
            "Value": aggregate.get("recommended_next_repair_target", ""),
        },
    ]
    for issue_type, count in (aggregate.get("issue_type_counts", {}) or {}).items():
        rows.append({"Metric": f"issue_type:{issue_type}", "Value": count})
    return rows


def build_dispatcher_review_v3_rows_by_sheet(
    dispatcher_rows,
    audit_rows,
    include_private_values=False,
    feedback_aggregate=None,
):
    return {
        SHEET_DISPATCHER_REVIEW: dispatcher_rows or [],
        SHEET_EXTRACTION_AUDIT: audit_rows or [],
        SHEET_REVIEW_INSTRUCTIONS: dispatcher_instruction_rows(
            include_private_values=include_private_values
        ),
        SHEET_FEEDBACK_SUMMARY: dispatcher_feedback_summary_rows(feedback_aggregate),
    }


def summarize_dispatcher_review_v3_rows(rows_by_sheet):
    return {
        "document_rows": len(
            (rows_by_sheet or {}).get(SHEET_DISPATCHER_REVIEW, []) or []
        ),
        "audit_rows": len((rows_by_sheet or {}).get(SHEET_EXTRACTION_AUDIT, []) or []),
        "instruction_rows": len(
            (rows_by_sheet or {}).get(SHEET_REVIEW_INSTRUCTIONS, []) or []
        ),
        "feedback_summary_rows": len(
            (rows_by_sheet or {}).get(SHEET_FEEDBACK_SUMMARY, []) or []
        ),
        "private_values_printed": False,
        "raw_text_printed": False,
        "money_values_printed": False,
    }


def _write_csv(path, rows, columns):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        for row in rows or []:
            writer.writerow({column: row.get(column, "") for column in columns})


def _write_xlsx_if_available(path, rows_by_sheet):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception:
        return False

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    header_fill = PatternFill("solid", fgColor="D9EAF7")

    for sheet_name, columns in DISPATCHER_REVIEW_COLUMNS_BY_SHEET.items():
        sheet = workbook.create_sheet(title=sheet_name)
        sheet.append(columns)
        for row in (rows_by_sheet or {}).get(sheet_name, []) or []:
            sheet.append([row.get(column, "") for column in columns])
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for index, column in enumerate(columns, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = min(
                max(len(column) + 2, 12),
                44,
            )

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return True


def write_dispatcher_review_v3_artifacts(
    dispatcher_rows,
    audit_rows,
    output_dir=None,
    include_private_values=False,
    feedback_aggregate=None,
    write_workbook=True,
    write_csvs=True,
    allow_custom_output_dir=False,
):
    output_root = _normalize_output_dir(
        output_dir or DEFAULT_PRIVATE_MEASUREMENT_OUTPUT_DIR,
        allow_custom_output_dir=allow_custom_output_dir,
    )
    rows_by_sheet = build_dispatcher_review_v3_rows_by_sheet(
        dispatcher_rows,
        audit_rows,
        include_private_values=include_private_values,
        feedback_aggregate=feedback_aggregate,
    )
    paths = {}
    if write_csvs:
        csv_specs = {
            "dispatcher_review_csv": (
                DISPATCHER_REVIEW_V3_REVIEW_CSV,
                SHEET_DISPATCHER_REVIEW,
                DISPATCHER_REVIEW_COLUMNS,
            ),
            "extraction_audit_csv": (
                DISPATCHER_REVIEW_V3_AUDIT_CSV,
                SHEET_EXTRACTION_AUDIT,
                DISPATCHER_AUDIT_COLUMNS,
            ),
            "instructions_csv": (
                DISPATCHER_REVIEW_V3_INSTRUCTIONS_CSV,
                SHEET_REVIEW_INSTRUCTIONS,
                DISPATCHER_INSTRUCTIONS_COLUMNS,
            ),
            "feedback_summary_csv": (
                DISPATCHER_REVIEW_V3_FEEDBACK_SUMMARY_CSV,
                SHEET_FEEDBACK_SUMMARY,
                DISPATCHER_FEEDBACK_SUMMARY_COLUMNS,
            ),
        }
        for key, (filename, sheet_name, columns) in csv_specs.items():
            path = output_root / filename
            _write_csv(path, rows_by_sheet.get(sheet_name, []), columns)
            paths[key] = path
    xlsx_written = False
    if write_workbook:
        xlsx_path = output_root / DISPATCHER_REVIEW_V3_WORKBOOK_XLSX
        xlsx_written = _write_xlsx_if_available(xlsx_path, rows_by_sheet)
        if xlsx_written:
            paths["dispatcher_review_workbook_xlsx"] = xlsx_path
    return {
        "paths": paths,
        "rows_by_sheet": rows_by_sheet,
        "summary": summarize_dispatcher_review_v3_rows(rows_by_sheet),
        "xlsx_written": xlsx_written,
        "csvs_written": bool(write_csvs),
        "include_private_values_local_only": bool(include_private_values),
        "local_only": True,
        "private_values_printed": False,
        "raw_text_printed": False,
        "money_values_printed": False,
    }
