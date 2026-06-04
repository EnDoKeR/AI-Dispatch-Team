"""Local-only RateCon review workbook row contracts and writers.

The workbook rows are designed for ignored local review artifacts. Console and
shareable summaries should only use the count metadata returned by this module.
"""

import csv
from pathlib import Path

from app.document_ai.extraction_readiness import assess_extraction_readiness
from app.document_ai.measurement_integrity import (
    check_measurement_row_integrity,
    summarize_integrity_issues,
)
from app.document_ai.private_measurement_outputs import (
    _normalize_output_dir,
)
from app.document_ai.measurement_cli.ratecon_private_output_paths import (
    REVIEW_DOCUMENT_SUMMARY_CSV,
    REVIEW_FIELD_REVIEW_CSV,
    REVIEW_RATE_REVIEW_CSV,
    REVIEW_STOP_REVIEW_CSV,
    REVIEW_V2_CORE_FIELDS_CSV,
    REVIEW_V2_DOCUMENT_SUMMARY_CSV,
    REVIEW_V2_INSTRUCTIONS_CSV,
    REVIEW_V2_LOAD_IDS_CSV,
    REVIEW_V2_RATES_CSV,
    REVIEW_V2_STOPS_CSV,
    REVIEW_V2_WORKBOOK_XLSX,
    REVIEW_WORKBOOK_XLSX,
    review_document_summary_csv_path,
    review_field_review_csv_path,
    review_rate_review_csv_path,
    review_stop_review_csv_path,
    review_v2_core_fields_csv_path,
    review_v2_document_summary_csv_path,
    review_v2_instructions_csv_path,
    review_v2_load_ids_csv_path,
    review_v2_rates_csv_path,
    review_v2_stops_csv_path,
    review_v2_workbook_path,
    review_workbook_path,
)
from app.document_ai.ratecon_candidates import normalize_list
from app.document_ai.ratecon_core_field_policy import (
    FIELD_POLICY_ROLE_DISPATCH_DECISION,
    FIELD_POLICY_ROLE_INTAKE_CORE,
    FIELD_REQUIREMENT_NON_APPLICABLE,
    FIELD_REQUIREMENT_OPTIONAL,
    FIELD_REQUIREMENT_REVIEW_REQUIRED,
    build_document_context,
    classify_field_policy_gap,
    get_field_requirement,
    is_field_blocker_for_level,
)


SHEET_DOCUMENT_SUMMARY = "Document_Summary"
SHEET_STOP_REVIEW = "Stop_Review"
SHEET_FIELD_REVIEW = "Field_Review"
SHEET_RATE_REVIEW = "Rate_Review"
SHEET_INSTRUCTIONS = "Instructions"

SHEET_V2_INSTRUCTIONS = "Review_Instructions"
SHEET_V2_DOCUMENT_SUMMARY = "Document_Summary"
SHEET_V2_CORE_FIELDS = "Core_Field_Review"
SHEET_V2_STOPS = "Stop_Review"
SHEET_V2_RATES = "Rate_Review"
SHEET_V2_LOAD_IDS = "Load_ID_Review"

DOCUMENT_SUMMARY_COLUMNS = [
    "Folder Order",
    "Local Document Name / File Stem",
    "Measurement Alias",
    "Document Type",
    "Classification Status",
    "Extraction Relevant",
    "Normal Load Movement",
    "TONU",
    "OCR Needed",
    "Layout Attempted",
    "Provider Status",
    "Old Raw Stop Groups",
    "Old Normalized Stops",
    "Span Anchors",
    "Stop Spans",
    "Span Normalized Stops",
    "Pickup Count",
    "Delivery Count",
    "Generic Stop Count",
    "Unknown Count",
    "Date Resolved",
    "Date Missing",
    "Time Resolved",
    "Time Missing",
    "Review Required Stops",
    "Readiness Level",
    "Extraction Review Blockers",
    "Intake Core Blockers",
    "Dispatch Decision Blockers",
    "Optional Missing Fields",
    "Non Applicable Fields",
    "Top Blocker",
    "Integrity Issues",
    "Recommended Review Priority",
    "User Review Status",
    "User Notes Local Only",
]

STOP_REVIEW_COLUMNS = [
    "Folder Order",
    "Local Document Name / File Stem",
    "Measurement Alias",
    "Stop ID",
    "Stop Sequence",
    "Stop Type",
    "Field Name",
    "Predicted Value LOCAL ONLY",
    "Status",
    "Confidence Bucket",
    "Evidence Type",
    "Page Number",
    "Source",
    "Needs Review",
    "Integrity Issue",
    "User Correct? yes/no/unknown",
    "User Expected Value LOCAL ONLY",
    "User Issue Type",
    "User Notes Local Only",
]

FIELD_REVIEW_COLUMNS = [
    "Folder Order",
    "Local Document Name / File Stem",
    "Measurement Alias",
    "Field Name",
    "Predicted Value LOCAL ONLY",
    "Status",
    "Confidence Bucket",
    "Evidence Type",
    "Needs Review",
    "Extraction Review Blocker",
    "Intake Core Blocker",
    "Dispatch Decision Blocker",
    "Review Field",
    "Optional Missing Field",
    "Non Applicable Field",
    "Field Requirement Level",
    "Policy Gap Reason",
    "Load Identifier Status",
    "Load Identifier Candidate Count",
    "Primary Load Identifier Candidate Type",
    "Typed Reference Count",
    "Rejected Non-primary Reference Count",
    "Load Identifier Gap Reason",
    "Load Identifier Needs Review",
    "User Correct? yes/no/unknown",
    "User Expected Value LOCAL ONLY",
    "User Issue Type",
    "User Notes Local Only",
]

RATE_REVIEW_COLUMNS = [
    "Measurement Alias",
    "Rate Field Type",
    "Predicted Value LOCAL ONLY",
    "Status",
    "Evidence Type",
    "Main Rate? yes/no/unknown",
    "Rate Conflict Reason",
    "Main Rate Candidate Count",
    "Equivalent Rate Group Count",
    "Different Strong Total Count",
    "Rate Review Required Reason",
    "Rate Selected? yes/no",
    "Rate Core Mapped? yes/no",
    "Rate Component Type",
    "User Correct? yes/no/unknown",
]

INSTRUCTIONS_COLUMNS = ["Section", "Instruction"]

V2_INSTRUCTIONS_COLUMNS = ["Section", "Instruction"]

V2_DOCUMENT_SUMMARY_COLUMNS = [
    "Folder Order",
    "Local Document Name / File Stem",
    "Measurement Alias",
    "Document Type",
    "OCR Needed",
    "Extraction Relevant",
    "Readiness Level",
    "Top Blockers",
    "Review Priority",
    "User Document Type Correct?",
    "User Notes Local Only",
]

V2_CORE_FIELD_COLUMNS = [
    "Measurement Alias",
    "Field Name",
    "Predicted Value LOCAL ONLY",
    "Predicted Status",
    "Needs Review",
    "Candidate Count",
    "Gap Reason",
    "Evidence Type",
    "User Correct? yes/no/unknown",
    "User Expected Value LOCAL ONLY",
    "User Issue Type",
    "User Notes Local Only",
]

V2_STOP_COLUMNS = [
    "Measurement Alias",
    "Stop ID",
    "Stop Type",
    "Sequence",
    "Field Name",
    "Predicted Value LOCAL ONLY",
    "Status",
    "User Correct? yes/no/unknown",
    "User Expected Value LOCAL ONLY",
    "User Issue Type",
]

V2_RATE_COLUMNS = [
    "Measurement Alias",
    "Rate Candidate Type",
    "Predicted Value LOCAL ONLY",
    "Status",
    "Conflict Reason",
    "Main Rate? yes/no/unknown",
    "User Correct? yes/no/unknown",
    "User Issue Type",
]

V2_LOAD_ID_COLUMNS = [
    "Measurement Alias",
    "Identifier Type",
    "Predicted Value LOCAL ONLY",
    "Primary Candidate?",
    "Rejected Non-primary?",
    "User Correct? yes/no/unknown",
    "User Issue Type",
]

REVIEW_WORKBOOK_COLUMNS_BY_SHEET = {
    SHEET_DOCUMENT_SUMMARY: DOCUMENT_SUMMARY_COLUMNS,
    SHEET_STOP_REVIEW: STOP_REVIEW_COLUMNS,
    SHEET_FIELD_REVIEW: FIELD_REVIEW_COLUMNS,
    SHEET_RATE_REVIEW: RATE_REVIEW_COLUMNS,
    SHEET_INSTRUCTIONS: INSTRUCTIONS_COLUMNS,
}

REVIEW_V2_COLUMNS_BY_SHEET = {
    SHEET_V2_INSTRUCTIONS: V2_INSTRUCTIONS_COLUMNS,
    SHEET_V2_DOCUMENT_SUMMARY: V2_DOCUMENT_SUMMARY_COLUMNS,
    SHEET_V2_CORE_FIELDS: V2_CORE_FIELD_COLUMNS,
    SHEET_V2_STOPS: V2_STOP_COLUMNS,
    SHEET_V2_RATES: V2_RATE_COLUMNS,
    SHEET_V2_LOAD_IDS: V2_LOAD_ID_COLUMNS,
}

LOCAL_PRIVATE_REVIEW_WARNING = (
    "LOCAL PRIVATE REVIEW ONLY - DO NOT COMMIT - DO NOT PASTE INTO CHAT"
)

RATE_REVIEW_FIELD_NAMES = {
    "rate",
    "payment_amount",
    "linehaul",
    "fuel_surcharge",
    "accessorial",
    "detention",
    "tonu_payment",
}

V2_CORE_FIELD_NAMES = {
    "broker_name",
    "load_number",
    "rate",
    "pickup_location",
    "pickup_date",
    "delivery_location",
    "delivery_date",
}

V2_STOP_FIELD_NAMES = {
    "location",
    "date",
    "time",
    "pickup_location",
    "pickup_date",
    "pickup_time",
    "delivery_location",
    "delivery_date",
    "delivery_time",
}


def _text(value):
    return str(value or "").strip()


def _token(value):
    return _text(value).lower().replace(" ", "_").replace("-", "_")


def _int(value):
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def _bool_text(value):
    return "yes" if bool(value) else "no"


def _join(values):
    return ";".join(normalize_list(values))


def _format_count_map(value):
    if not isinstance(value, dict):
        return ""
    return ";".join(
        f"{_token(key)}={_int(count)}"
        for key, count in sorted(value.items())
        if _token(key) and _int(count)
    )


def _selected_value(item, include_private_values=False):
    if not include_private_values or not isinstance(item, dict):
        return ""
    for key in (
        "selected_value",
        "predicted_value",
        "value",
        "extracted_value",
        "value_text",
    ):
        value = _text(item.get(key))
        if value:
            return value
    return ""


def _first_evidence_ref(field):
    refs = (field or {}).get("evidence_refs", []) if isinstance(field, dict) else []
    for ref in refs:
        if isinstance(ref, dict):
            return ref
    return {}


def _evidence_type(item):
    if not isinstance(item, dict):
        return ""
    if _text(item.get("evidence_type")):
        return _text(item.get("evidence_type"))
    ref = _first_evidence_ref(item)
    return _text(ref.get("evidence_type"))


def _page_number(item):
    if not isinstance(item, dict):
        return ""
    if _text(item.get("page_number")):
        return _text(item.get("page_number"))
    ref = _first_evidence_ref(item)
    return _text(ref.get("page_number"))


def _field_status(row, field_name):
    for field_status in (row or {}).get("field_statuses", []) or []:
        if isinstance(field_status, dict) and _token(field_status.get("field_name")) == field_name:
            return _text(field_status.get("status"))
    return ""


def _field_status_map(row):
    statuses = {}
    for field_status in (row or {}).get("field_statuses", []) or []:
        if not isinstance(field_status, dict):
            continue
        field_name = _token(field_status.get("field_name"))
        if field_name:
            statuses[field_name] = _token(field_status.get("status"))
    for field_name in normalize_list((row or {}).get("missing_fields", [])):
        statuses.setdefault(_token(field_name), "missing")
    for field_name in normalize_list((row or {}).get("needs_check_fields", [])):
        statuses.setdefault(_token(field_name), "needs_review")
    for field_name in normalize_list((row or {}).get("conflict_fields", [])):
        statuses[_token(field_name)] = "conflict"
    for field_name in normalize_list((row or {}).get("non_applicable_fields", [])):
        statuses[_token(field_name)] = "non_applicable"
    return {key: value for key, value in statuses.items() if key}


def _field_policy_columns(row, field_name, status):
    statuses = _field_status_map(row)
    context = build_document_context(row, field_statuses=statuses)
    requirement = get_field_requirement(
        field_name,
        FIELD_POLICY_ROLE_INTAKE_CORE,
        context,
    )
    intake_blocker = is_field_blocker_for_level(
        field_name,
        status,
        FIELD_POLICY_ROLE_INTAKE_CORE,
        context,
    )
    dispatch_blocker = is_field_blocker_for_level(
        field_name,
        status,
        FIELD_POLICY_ROLE_DISPATCH_DECISION,
        context,
    )
    optional_missing = (
        _token(status) == "missing"
        and requirement
        in {
            FIELD_REQUIREMENT_OPTIONAL,
            FIELD_REQUIREMENT_REVIEW_REQUIRED,
        }
    )
    non_applicable = requirement == FIELD_REQUIREMENT_NON_APPLICABLE
    review_field = (
        requirement
        in {
            FIELD_REQUIREMENT_OPTIONAL,
            FIELD_REQUIREMENT_REVIEW_REQUIRED,
        }
        or dispatch_blocker
    ) and not non_applicable
    extraction_review_blocker = (
        _text((row or {}).get("extraction_status")).upper() == "EMPTY_TEXT"
        and _token(status) == "missing"
    )
    return {
        "Extraction Review Blocker": _bool_text(extraction_review_blocker),
        "Intake Core Blocker": _bool_text(intake_blocker),
        "Dispatch Decision Blocker": _bool_text(dispatch_blocker),
        "Review Field": _bool_text(review_field),
        "Optional Missing Field": _bool_text(optional_missing),
        "Non Applicable Field": _bool_text(non_applicable),
        "Field Requirement Level": requirement,
        "Policy Gap Reason": classify_field_policy_gap(
            field_name,
            status,
            FIELD_POLICY_ROLE_INTAKE_CORE,
            context,
        ),
    }


def _load_identifier_gap_reason(metrics):
    if not isinstance(metrics, dict):
        return ""
    identifier_label_count = _int(metrics.get("identifier_label_feature_count"))
    primary_count = _int(metrics.get("primary_identifier_candidate_count"))
    typed_reference_count = _int(metrics.get("typed_reference_candidate_count"))
    core_mapping_count = _int(metrics.get("core_load_number_mapping_count"))
    if primary_count == 0:
        if typed_reference_count:
            return "only_non_primary_reference_found"
        if identifier_label_count == 0:
            return "identifier_label_missing"
        return "identifier_candidate_not_generated"
    if core_mapping_count == 0:
        if _int(metrics.get("conflicting_primary_identifiers")):
            return "conflicting_primary_identifiers"
        if _int(metrics.get("weak_generic_reference_review_required")):
            return "weak_generic_reference_review_required"
        return "candidate_generated_but_not_core_mapped"
    return ""


def _load_identifier_review_columns(row, field_name, status):
    if _token(field_name) != "load_number":
        return {
            "Load Identifier Status": "",
            "Load Identifier Candidate Count": "",
            "Primary Load Identifier Candidate Type": "",
            "Typed Reference Count": "",
            "Rejected Non-primary Reference Count": "",
            "Load Identifier Gap Reason": "",
            "Load Identifier Needs Review": "",
        }
    metrics = (row or {}).get("load_identifier_coverage_metrics", {}) or {}
    primary_count = _int(metrics.get("primary_identifier_candidate_count"))
    return {
        "Load Identifier Status": _text(status),
        "Load Identifier Candidate Count": primary_count,
        "Primary Load Identifier Candidate Type": _format_count_map(
            metrics.get("primary_identifier_type_counts")
        ),
        "Typed Reference Count": _int(metrics.get("typed_reference_candidate_count")),
        "Rejected Non-primary Reference Count": _int(
            metrics.get("rejected_reference_as_load_id_count")
        ),
        "Load Identifier Gap Reason": _load_identifier_gap_reason(metrics),
        "Load Identifier Needs Review": _bool_text(
            _token(status) in {"missing", "needs_review", "conflict", "low_confidence"}
        ),
    }


def _top_blocker(row):
    blockers = normalize_list((row or {}).get("blocker_categories", []))
    return blockers[0] if blockers else ""


def _layout_attempted(row):
    status = _text((row or {}).get("layout_provider_status"))
    return bool(status) and not status.startswith("skipped")


def _review_priority(row, integrity_issues, readiness):
    if integrity_issues:
        return "integrity_check"
    if readiness.get("readiness_level") == "not_ready":
        return "high"
    if _int((row or {}).get("span_review_required_count")):
        return "stop_review"
    if _int((row or {}).get("old_normalized_stops")) > _int(
        (row or {}).get("span_normalized_stop_count")
    ):
        return "value_review"
    return "normal"


def _stop_set_for_review(row):
    span_stop_set = (row or {}).get("span_normalized_stop_set", {}) or {}
    if span_stop_set.get("stops"):
        return span_stop_set, "span_normalized_stop_set"
    return (row or {}).get("normalized_stop_set", {}) or {}, "normalized_stop_set"


def _stop_needs_review(stop, field):
    return bool((stop or {}).get("review_required")) or _text((field or {}).get("status")) in {
        "missing",
        "low_confidence",
        "conflict",
        "review_required",
        "needs_review",
    }


def _integrity_issue_codes(row):
    return [issue.get("issue_code", "") for issue in check_measurement_row_integrity(row)]


def _field_review_rows(row, folder_order, local_name, include_private_values=False):
    rows = []
    alias = _text((row or {}).get("document_alias"))
    seen = set()
    for field in (row or {}).get("field_statuses", []) or []:
        if not isinstance(field, dict):
            continue
        field_name = _token(field.get("field_name"))
        if not field_name:
            continue
        seen.add(field_name)
        status = _text(field.get("status"))
        policy_columns = _field_policy_columns(row, field_name, status)
        rows.append(
            {
                "Folder Order": folder_order,
                "Local Document Name / File Stem": local_name,
                "Measurement Alias": alias,
                "Field Name": field_name,
                "Predicted Value LOCAL ONLY": _selected_value(
                    field,
                    include_private_values=include_private_values,
                ),
                "Status": status,
                "Confidence Bucket": _text(field.get("confidence")),
                "Evidence Type": _evidence_type(field),
                "Needs Review": _bool_text(
                    status
                    in {"missing", "low_confidence", "conflict", "needs_review"}
                ),
                **policy_columns,
                **_load_identifier_review_columns(row, field_name, status),
                "User Correct? yes/no/unknown": "",
                "User Expected Value LOCAL ONLY": "",
                "User Issue Type": "",
                "User Notes Local Only": "",
            }
        )

    for field_name in normalize_list((row or {}).get("missing_fields", [])):
        token = _token(field_name)
        if token and token not in seen:
            rows.append(
                {
                    "Folder Order": folder_order,
                    "Local Document Name / File Stem": local_name,
                    "Measurement Alias": alias,
                    "Field Name": token,
                    "Predicted Value LOCAL ONLY": "",
                    "Status": "missing",
                    "Confidence Bucket": "",
                    "Evidence Type": "",
                    "Needs Review": "yes",
                    **_field_policy_columns(row, token, "missing"),
                    **_load_identifier_review_columns(row, token, "missing"),
                    "User Correct? yes/no/unknown": "",
                    "User Expected Value LOCAL ONLY": "",
                    "User Issue Type": "",
                    "User Notes Local Only": "",
                }
            )
    return rows


def _first_rate_conflict_audit_record(row):
    for record in (row or {}).get("rate_conflict_audit_records", []) or []:
        if isinstance(record, dict):
            return record
    return {}


def _first_rate_forensics_record(row):
    for record in (row or {}).get("rate_forensics_records", []) or []:
        if isinstance(record, dict):
            return record
    return {}


def _rate_review_rows(row, field_rows):
    rows = []
    conflict_record = _first_rate_conflict_audit_record(row)
    forensics_record = _first_rate_forensics_record(row)
    conflict_reason = _token(conflict_record.get("conflict_reason"))
    review_required = bool(conflict_record.get("review_required"))
    for field_row in field_rows or []:
        if _token(field_row.get("Field Name")) not in RATE_REVIEW_FIELD_NAMES:
            continue
        rows.append(
            {
                "Measurement Alias": field_row.get("Measurement Alias", ""),
                "Rate Field Type": field_row.get("Field Name", ""),
                "Predicted Value LOCAL ONLY": field_row.get(
                    "Predicted Value LOCAL ONLY",
                    "",
                ),
                "Status": field_row.get("Status", ""),
                "Evidence Type": field_row.get("Evidence Type", ""),
                "Main Rate? yes/no/unknown": "",
                "Rate Conflict Reason": conflict_reason,
                "Main Rate Candidate Count": _int(
                    conflict_record.get("main_rate_candidate_count")
                    or forensics_record.get("main_rate_candidate_count")
                ),
                "Equivalent Rate Group Count": _int(
                    conflict_record.get("equivalent_candidate_group_count")
                ),
                "Different Strong Total Count": _int(
                    conflict_record.get("different_strong_total_count")
                ),
                "Rate Review Required Reason": conflict_reason if review_required else "",
                "Rate Selected? yes/no": _bool_text(
                    conflict_record.get("selected_rate_present")
                ),
                "Rate Core Mapped? yes/no": _bool_text(
                    conflict_record.get("core_rate_mapped")
                ),
                "Rate Component Type": _format_count_map(
                    forensics_record.get("category_counts")
                ),
                "User Correct? yes/no/unknown": "",
            }
        )
    return rows


def _stop_review_rows(row, folder_order, local_name, include_private_values=False):
    stop_set, source = _stop_set_for_review(row)
    alias = _text((row or {}).get("document_alias"))
    integrity_codes = _join(_integrity_issue_codes(row))
    rows = []
    for stop in stop_set.get("stops", []) or []:
        if not isinstance(stop, dict):
            continue
        for field in stop.get("fields", []) or []:
            if not isinstance(field, dict):
                continue
            rows.append(
                {
                    "Folder Order": folder_order,
                    "Local Document Name / File Stem": local_name,
                    "Measurement Alias": alias,
                    "Stop ID": _text(stop.get("stop_id")),
                    "Stop Sequence": _text(stop.get("sequence")),
                    "Stop Type": _text(stop.get("stop_type")),
                    "Field Name": _text(field.get("field_name")),
                    "Predicted Value LOCAL ONLY": _selected_value(
                        field,
                        include_private_values=include_private_values,
                    ),
                    "Status": _text(field.get("status")),
                    "Confidence Bucket": _text(field.get("confidence")),
                    "Evidence Type": _evidence_type(field),
                    "Page Number": _page_number(field),
                    "Source": source,
                    "Needs Review": _bool_text(_stop_needs_review(stop, field)),
                    "Integrity Issue": integrity_codes,
                    "User Correct? yes/no/unknown": "",
                    "User Expected Value LOCAL ONLY": "",
                    "User Issue Type": "",
                    "User Notes Local Only": "",
                }
            )
    return rows


def _document_summary_row(row, folder_order, local_name):
    readiness = assess_extraction_readiness(row)
    issues = check_measurement_row_integrity(row)
    issue_codes = [issue.get("issue_code", "") for issue in issues]
    return {
        "Folder Order": folder_order,
        "Local Document Name / File Stem": local_name,
        "Measurement Alias": _text((row or {}).get("document_alias")),
        "Document Type": _text((row or {}).get("document_type")),
        "Classification Status": _text((row or {}).get("classification_status")),
        "Extraction Relevant": _bool_text((row or {}).get("extraction_relevant")),
        "Normal Load Movement": _bool_text((row or {}).get("normal_load_movement")),
        "TONU": _bool_text(_text((row or {}).get("document_type")).upper() == "TRUCK_ORDER_NOT_USED"),
        "OCR Needed": _bool_text((row or {}).get("extraction_status") == "EMPTY_TEXT"),
        "Layout Attempted": _bool_text(_layout_attempted(row)),
        "Provider Status": _text((row or {}).get("layout_provider_status")),
        "Old Raw Stop Groups": _int((row or {}).get("old_raw_stop_groups", (row or {}).get("raw_stop_group_count"))),
        "Old Normalized Stops": _int((row or {}).get("old_normalized_stops", (row or {}).get("normalized_stop_count"))),
        "Span Anchors": _int((row or {}).get("span_anchor_count")),
        "Stop Spans": _int((row or {}).get("stop_span_count")),
        "Span Normalized Stops": _int((row or {}).get("span_normalized_stop_count")),
        "Pickup Count": _int((row or {}).get("span_pickup_count", (row or {}).get("pickup_count"))),
        "Delivery Count": _int((row or {}).get("span_delivery_count", (row or {}).get("delivery_count"))),
        "Generic Stop Count": _int(
            (row or {}).get("span_generic_stop_count", (row or {}).get("generic_stop_count"))
        ),
        "Unknown Count": _int((row or {}).get("span_unknown_count", (row or {}).get("unknown_stop_count"))),
        "Date Resolved": _int((row or {}).get("span_date_resolved_count")),
        "Date Missing": _int((row or {}).get("span_date_missing_count")),
        "Time Resolved": _int((row or {}).get("span_time_resolved_count")),
        "Time Missing": _int((row or {}).get("span_time_missing_count")),
        "Review Required Stops": _int((row or {}).get("span_review_required_count")),
        "Readiness Level": readiness.get("readiness_level", "not_ready"),
        "Extraction Review Blockers": _join(readiness.get("extraction_review_blockers")),
        "Intake Core Blockers": _join(readiness.get("intake_core_blockers")),
        "Dispatch Decision Blockers": _join(readiness.get("dispatch_decision_blockers")),
        "Optional Missing Fields": _join(readiness.get("optional_missing_fields")),
        "Non Applicable Fields": _join(readiness.get("non_applicable_fields")),
        "Top Blocker": _top_blocker(row),
        "Integrity Issues": _join(issue_codes),
        "Recommended Review Priority": _review_priority(row, issues, readiness),
        "User Review Status": "",
        "User Notes Local Only": "",
    }


def _instruction_rows(include_private_values=False):
    rows = [
        {
            "Section": "Local-only warning",
            "Instruction": LOCAL_PRIVATE_REVIEW_WARNING,
        },
        {
            "Section": "Review order",
            "Instruction": "Review Document_Summary first, then Stop_Review, Field_Review, and Rate_Review.",
        },
        {
            "Section": "Safe sharing",
            "Instruction": "Share aliases, counts, statuses, issue types, and field names only.",
        },
        {
            "Section": "Do not share",
            "Instruction": "Do not share private values, raw text, local paths, broker names, rates, addresses, or references.",
        },
    ]
    if include_private_values:
        rows.append(
            {
                "Section": "Private values",
                "Instruction": "Predicted values are included only for local review. Do not paste workbook rows into chat.",
            }
        )
    else:
        rows.append(
            {
                "Section": "Private values",
                "Instruction": "Predicted value columns are intentionally blank in redacted mode.",
            }
        )
    return rows


def build_ratecon_review_rows(
    measurement_rows,
    local_document_names_by_alias=None,
    include_private_values=False,
):
    local_names = local_document_names_by_alias or {}
    document_rows = []
    stop_rows = []
    field_rows = []
    rate_rows = []

    for folder_order, row in enumerate(measurement_rows or [], start=1):
        if not isinstance(row, dict):
            continue
        alias = _text(row.get("document_alias"))
        local_name = _text(local_names.get(alias))
        document_rows.append(_document_summary_row(row, folder_order, local_name))
        row_stop_rows = _stop_review_rows(
            row,
            folder_order,
            local_name,
            include_private_values=include_private_values,
        )
        row_field_rows = _field_review_rows(
            row,
            folder_order,
            local_name,
            include_private_values=include_private_values,
        )
        stop_rows.extend(row_stop_rows)
        field_rows.extend(row_field_rows)
        rate_rows.extend(_rate_review_rows(row, row_field_rows))

    return {
        SHEET_DOCUMENT_SUMMARY: document_rows,
        SHEET_STOP_REVIEW: stop_rows,
        SHEET_FIELD_REVIEW: field_rows,
        SHEET_RATE_REVIEW: rate_rows,
        SHEET_INSTRUCTIONS: _instruction_rows(
            include_private_values=include_private_values
        ),
    }


def _v2_instruction_rows(include_private_values=False):
    rows = [
        {
            "Section": "Review order",
            "Instruction": "Review Document_Summary, Core_Field_Review, Stop_Review, Rate_Review, then Load_ID_Review.",
        },
        {
            "Section": "Review columns",
            "Instruction": "Fill User Correct?, User Issue Type, and User Expected Value LOCAL ONLY when needed.",
        },
        {
            "Section": "Safe sharing",
            "Instruction": "Share only aliases, counts, statuses, issue types, field names, and readiness counts.",
        },
        {
            "Section": "Do not share",
            "Instruction": "Do not share private values, expected values, raw text, local paths, filenames, service account keys, or money amounts.",
        },
    ]
    rows.append(
        {
            "Section": "Predicted values",
            "Instruction": (
                "Predicted values are included for local-private review only."
                if include_private_values
                else "Predicted value columns are blank in status-only mode."
            ),
        }
    )
    return rows


def _v2_document_summary_row(row, folder_order, local_name):
    readiness = assess_extraction_readiness(row)
    top_blockers = normalize_list(readiness.get("intake_core_blockers"))
    if not top_blockers:
        top_blockers = normalize_list(readiness.get("extraction_review_blockers"))
    if not top_blockers:
        top_blockers = normalize_list(readiness.get("dispatch_decision_blockers"))
    issues = check_measurement_row_integrity(row)
    return {
        "Folder Order": folder_order,
        "Local Document Name / File Stem": local_name,
        "Measurement Alias": _text((row or {}).get("document_alias")),
        "Document Type": _text((row or {}).get("document_type")),
        "OCR Needed": _bool_text((row or {}).get("extraction_status") == "EMPTY_TEXT"),
        "Extraction Relevant": _bool_text((row or {}).get("extraction_relevant")),
        "Readiness Level": readiness.get("readiness_level", "not_ready"),
        "Top Blockers": _join(top_blockers[:5]),
        "Review Priority": _review_priority(row, issues, readiness),
        "User Document Type Correct?": "",
        "User Notes Local Only": "",
    }


def _v2_candidate_count(row, field_name):
    metrics = (row or {}).get("candidate_coverage_metrics", {}) or {}
    field_counts = metrics.get("span_field_candidate_counts_by_field", {})
    if isinstance(field_counts, dict) and field_name in field_counts:
        return _int(field_counts.get(field_name))
    if field_name == "load_number":
        return _int(
            ((row or {}).get("load_identifier_coverage_metrics", {}) or {}).get(
                "primary_identifier_candidate_count"
            )
        )
    if field_name == "rate":
        for record in (row or {}).get("rate_forensics_records", []) or []:
            if isinstance(record, dict):
                return _int(record.get("main_rate_candidate_count"))
    return ""


def _v2_gap_reason(row, field_row):
    field_name = _token((field_row or {}).get("Field Name"))
    if field_name == "load_number":
        reason = _text((field_row or {}).get("Load Identifier Gap Reason"))
        if reason:
            return reason
    if field_name == "rate":
        reason = _text((field_row or {}).get("Rate Conflict Reason"))
        if reason:
            return reason
    return _text((field_row or {}).get("Policy Gap Reason"))


def _v2_core_field_rows(
    row,
    folder_order,
    local_name,
    include_private_values=False,
    field_rows=None,
):
    field_rows = field_rows or _field_review_rows(
        row,
        folder_order,
        local_name,
        include_private_values=include_private_values,
    )
    rows = []
    seen = set()
    for field_row in field_rows:
        field_name = _token(field_row.get("Field Name"))
        if field_name not in V2_CORE_FIELD_NAMES:
            continue
        seen.add(field_name)
        rows.append(
            {
                "Measurement Alias": field_row.get("Measurement Alias", ""),
                "Field Name": field_name,
                "Predicted Value LOCAL ONLY": field_row.get(
                    "Predicted Value LOCAL ONLY",
                    "",
                ),
                "Predicted Status": field_row.get("Status", ""),
                "Needs Review": field_row.get("Needs Review", ""),
                "Candidate Count": _v2_candidate_count(row, field_name),
                "Gap Reason": _v2_gap_reason(row, field_row),
                "Evidence Type": field_row.get("Evidence Type", ""),
                "User Correct? yes/no/unknown": "",
                "User Expected Value LOCAL ONLY": "",
                "User Issue Type": "",
                "User Notes Local Only": "",
            }
        )
    alias = _text((row or {}).get("document_alias"))
    for field_name in sorted(V2_CORE_FIELD_NAMES - seen):
        rows.append(
            {
                "Measurement Alias": alias,
                "Field Name": field_name,
                "Predicted Value LOCAL ONLY": "",
                "Predicted Status": "missing",
                "Needs Review": "yes",
                "Candidate Count": _v2_candidate_count(row, field_name),
                "Gap Reason": classify_field_policy_gap(
                    field_name,
                    "missing",
                    FIELD_POLICY_ROLE_INTAKE_CORE,
                    build_document_context(row),
                ),
                "Evidence Type": "",
                "User Correct? yes/no/unknown": "",
                "User Expected Value LOCAL ONLY": "",
                "User Issue Type": "",
                "User Notes Local Only": "",
            }
        )
    return rows


def _v2_stop_rows(row, folder_order, local_name, include_private_values=False):
    rows = []
    for stop_row in _stop_review_rows(
        row,
        folder_order,
        local_name,
        include_private_values=include_private_values,
    ):
        stop_type = _token(stop_row.get("Stop Type"))
        field_name = _token(stop_row.get("Field Name"))
        if stop_type not in {"pickup", "delivery"}:
            continue
        if field_name not in V2_STOP_FIELD_NAMES and _token(stop_row.get("Needs Review")) != "yes":
            continue
        rows.append(
            {
                "Measurement Alias": stop_row.get("Measurement Alias", ""),
                "Stop ID": stop_row.get("Stop ID", ""),
                "Stop Type": stop_row.get("Stop Type", ""),
                "Sequence": stop_row.get("Stop Sequence", ""),
                "Field Name": stop_row.get("Field Name", ""),
                "Predicted Value LOCAL ONLY": stop_row.get(
                    "Predicted Value LOCAL ONLY",
                    "",
                ),
                "Status": stop_row.get("Status", ""),
                "User Correct? yes/no/unknown": "",
                "User Expected Value LOCAL ONLY": "",
                "User Issue Type": "",
            }
        )
    return rows


def _v2_rate_rows(row, field_rows):
    rate_rows = _rate_review_rows(row, field_rows)
    return [
        {
            "Measurement Alias": rate_row.get("Measurement Alias", ""),
            "Rate Candidate Type": rate_row.get("Rate Field Type", ""),
            "Predicted Value LOCAL ONLY": rate_row.get("Predicted Value LOCAL ONLY", ""),
            "Status": rate_row.get("Status", ""),
            "Conflict Reason": rate_row.get("Rate Conflict Reason", ""),
            "Main Rate? yes/no/unknown": rate_row.get("Main Rate? yes/no/unknown", ""),
            "User Correct? yes/no/unknown": "",
            "User Issue Type": "",
        }
        for rate_row in rate_rows
    ]


def _v2_load_identifier_rows(row, core_field_rows, include_private_values=False):
    alias = _text((row or {}).get("document_alias"))
    metrics = (row or {}).get("load_identifier_coverage_metrics", {}) or {}
    load_row = next(
        (field for field in core_field_rows if _token(field.get("Field Name")) == "load_number"),
        {},
    )
    rows = []
    primary_types = metrics.get("primary_identifier_type_counts", {}) or {}
    for identifier_type, count in sorted(primary_types.items()):
        for _index in range(_int(count) or 1):
            rows.append(
                {
                    "Measurement Alias": alias,
                    "Identifier Type": _token(identifier_type),
                    "Predicted Value LOCAL ONLY": load_row.get(
                        "Predicted Value LOCAL ONLY",
                        "",
                    )
                    if include_private_values
                    else "",
                    "Primary Candidate?": "yes",
                    "Rejected Non-primary?": "no",
                    "User Correct? yes/no/unknown": "",
                    "User Issue Type": "",
                }
            )
    rejected_count = _int(metrics.get("rejected_reference_as_load_id_count"))
    if rejected_count:
        rows.append(
            {
                "Measurement Alias": alias,
                "Identifier Type": "non_primary_reference",
                "Predicted Value LOCAL ONLY": "",
                "Primary Candidate?": "no",
                "Rejected Non-primary?": "yes",
                "User Correct? yes/no/unknown": "",
                "User Issue Type": "",
            }
        )
    if not rows and _token(load_row.get("Predicted Status")) in {
        "missing",
        "needs_review",
        "conflict",
        "low_confidence",
    }:
        rows.append(
            {
                "Measurement Alias": alias,
                "Identifier Type": "load_number",
                "Predicted Value LOCAL ONLY": "",
                "Primary Candidate?": "no",
                "Rejected Non-primary?": "no",
                "User Correct? yes/no/unknown": "",
                "User Issue Type": "",
            }
        )
    return rows


def build_ratecon_review_v2_rows(
    measurement_rows,
    local_document_names_by_alias=None,
    include_private_values=False,
):
    local_names = local_document_names_by_alias or {}
    document_rows = []
    core_rows = []
    stop_rows = []
    rate_rows = []
    load_id_rows = []

    for folder_order, row in enumerate(measurement_rows or [], start=1):
        if not isinstance(row, dict):
            continue
        alias = _text(row.get("document_alias"))
        local_name = _text(local_names.get(alias))
        document_rows.append(_v2_document_summary_row(row, folder_order, local_name))
        row_field_rows = _field_review_rows(
            row,
            folder_order,
            local_name,
            include_private_values=include_private_values,
        )
        row_core_rows = _v2_core_field_rows(
            row,
            folder_order,
            local_name,
            include_private_values=include_private_values,
            field_rows=row_field_rows,
        )
        core_rows.extend(row_core_rows)
        stop_rows.extend(
            _v2_stop_rows(
                row,
                folder_order,
                local_name,
                include_private_values=include_private_values,
            )
        )
        rate_rows.extend(_v2_rate_rows(row, row_field_rows))
        load_id_rows.extend(
            _v2_load_identifier_rows(
                row,
                row_core_rows,
                include_private_values=include_private_values,
            )
        )

    return {
        SHEET_V2_INSTRUCTIONS: _v2_instruction_rows(
            include_private_values=include_private_values
        ),
        SHEET_V2_DOCUMENT_SUMMARY: document_rows,
        SHEET_V2_CORE_FIELDS: core_rows,
        SHEET_V2_STOPS: stop_rows,
        SHEET_V2_RATES: rate_rows,
        SHEET_V2_LOAD_IDS: load_id_rows,
    }


def summarize_ratecon_review_v2_rows(rows_by_sheet):
    return {
        "document_rows": len(
            (rows_by_sheet or {}).get(SHEET_V2_DOCUMENT_SUMMARY, []) or []
        ),
        "core_field_rows": len(
            (rows_by_sheet or {}).get(SHEET_V2_CORE_FIELDS, []) or []
        ),
        "stop_rows": len((rows_by_sheet or {}).get(SHEET_V2_STOPS, []) or []),
        "rate_rows": len((rows_by_sheet or {}).get(SHEET_V2_RATES, []) or []),
        "load_id_rows": len((rows_by_sheet or {}).get(SHEET_V2_LOAD_IDS, []) or []),
        "instruction_rows": len(
            (rows_by_sheet or {}).get(SHEET_V2_INSTRUCTIONS, []) or []
        ),
        "raw_text_included": False,
        "private_values_printed": False,
    }


def summarize_ratecon_review_workbook_rows(rows_by_sheet):
    document_rows = (rows_by_sheet or {}).get(SHEET_DOCUMENT_SUMMARY, []) or []
    readiness_counts = {}
    for row in document_rows:
        level = _text(row.get("Readiness Level"))
        if level:
            readiness_counts[level] = readiness_counts.get(level, 0) + 1

    issue_codes = []
    for row in document_rows:
        issue_codes.extend(normalize_list(row.get("Integrity Issues", "").split(";")))
    integrity_summary = summarize_integrity_issues(
        [{"issue_code": code, "severity": "warning"} for code in issue_codes if code]
    )
    field_rows = (rows_by_sheet or {}).get(SHEET_FIELD_REVIEW, []) or []

    def count_field_rows(flag_column):
        counts = {}
        for row in field_rows:
            if _token(row.get(flag_column)) != "yes":
                continue
            field_name = _token(row.get("Field Name"))
            if field_name:
                counts[field_name] = counts.get(field_name, 0) + 1
        return dict(sorted(counts.items()))

    policy_misclassification_count = sum(
        1
        for row in field_rows
        if _token(row.get("Policy Gap Reason")) == "optional_field_misclassified_as_core"
    )
    return {
        "document_rows": len(document_rows),
        "stop_review_rows": len((rows_by_sheet or {}).get(SHEET_STOP_REVIEW, []) or []),
        "field_review_rows": len(field_rows),
        "rate_review_rows": len((rows_by_sheet or {}).get(SHEET_RATE_REVIEW, []) or []),
        "readiness_level_counts": dict(sorted(readiness_counts.items())),
        "integrity_issue_counts": integrity_summary.get("issue_counts", {}),
        "extraction_review_blocker_counts": count_field_rows("Extraction Review Blocker"),
        "intake_core_blocker_counts": count_field_rows("Intake Core Blocker"),
        "dispatch_decision_blocker_counts": count_field_rows("Dispatch Decision Blocker"),
        "optional_missing_field_counts": count_field_rows("Optional Missing Field"),
        "non_applicable_field_counts": count_field_rows("Non Applicable Field"),
        "policy_misclassification_count": policy_misclassification_count,
        "raw_text_included": False,
        "private_values_printed": False,
    }


def _write_csv(path, rows, columns):
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        for row in rows or []:
            writer.writerow({column: row.get(column, "") for column in columns})


def write_ratecon_review_csvs(
    measurement_rows,
    output_dir=None,
    local_document_names_by_alias=None,
    include_private_values=False,
    allow_custom_output_dir=False,
):
    output_root = _normalize_output_dir(output_dir, allow_custom_output_dir)
    rows_by_sheet = build_ratecon_review_rows(
        measurement_rows,
        local_document_names_by_alias=local_document_names_by_alias,
        include_private_values=include_private_values,
    )
    csv_specs = {
        "document_summary_csv": (
            review_document_summary_csv_path,
            SHEET_DOCUMENT_SUMMARY,
            DOCUMENT_SUMMARY_COLUMNS,
        ),
        "stop_review_csv": (
            review_stop_review_csv_path,
            SHEET_STOP_REVIEW,
            STOP_REVIEW_COLUMNS,
        ),
        "field_review_csv": (
            review_field_review_csv_path,
            SHEET_FIELD_REVIEW,
            FIELD_REVIEW_COLUMNS,
        ),
        "rate_review_csv": (
            review_rate_review_csv_path,
            SHEET_RATE_REVIEW,
            RATE_REVIEW_COLUMNS,
        ),
    }
    paths = {}
    for key, (path_builder, sheet_name, columns) in csv_specs.items():
        path = path_builder(output_root)
        _write_csv(path, rows_by_sheet.get(sheet_name, []), columns)
        paths[key] = path

    return {
        "paths": paths,
        "rows_by_sheet": rows_by_sheet,
        "summary": summarize_ratecon_review_workbook_rows(rows_by_sheet),
        "include_private_values_local_only": bool(include_private_values),
        "local_only": True,
        "raw_text_included": False,
        "private_values_printed": False,
    }


def _write_xlsx_with_columns_if_available(path, rows_by_sheet, columns_by_sheet):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception:
        return False

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    warning_fill = PatternFill("solid", fgColor="FDE68A")

    for sheet_name, columns in columns_by_sheet.items():
        sheet = workbook.create_sheet(title=sheet_name)
        sheet.append(columns)
        for row in rows_by_sheet.get(sheet_name, []) or []:
            sheet.append([row.get(column, "") for column in columns])

        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        if sheet_name == SHEET_INSTRUCTIONS:
            for cell in sheet[2]:
                cell.fill = warning_fill
                cell.font = Font(bold=True)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for index, column in enumerate(columns, start=1):
            width = min(max(len(column) + 2, 12), 42)
            sheet.column_dimensions[get_column_letter(index)].width = width

    workbook.save(path)
    return True


def _write_xlsx_if_available(path, rows_by_sheet):
    return _write_xlsx_with_columns_if_available(
        path,
        rows_by_sheet,
        REVIEW_WORKBOOK_COLUMNS_BY_SHEET,
    )


def _write_v2_xlsx_if_available(path, rows_by_sheet):
    return _write_xlsx_with_columns_if_available(
        path,
        rows_by_sheet,
        REVIEW_V2_COLUMNS_BY_SHEET,
    )


def write_ratecon_review_workbook(
    measurement_rows,
    output_dir=None,
    local_document_names_by_alias=None,
    include_private_values=False,
    allow_custom_output_dir=False,
):
    output_root = _normalize_output_dir(output_dir, allow_custom_output_dir)
    rows_by_sheet = build_ratecon_review_rows(
        measurement_rows,
        local_document_names_by_alias=local_document_names_by_alias,
        include_private_values=include_private_values,
    )
    xlsx_path = review_workbook_path(output_root)
    xlsx_written = _write_xlsx_if_available(xlsx_path, rows_by_sheet)
    return {
        "xlsx": xlsx_path if xlsx_written else None,
        "rows_by_sheet": rows_by_sheet,
        "summary": summarize_ratecon_review_workbook_rows(rows_by_sheet),
        "include_private_values_local_only": bool(include_private_values),
        "local_only": True,
        "raw_text_included": False,
        "private_values_printed": False,
    }


def write_ratecon_review_v2_artifacts(
    measurement_rows,
    output_dir=None,
    local_document_names_by_alias=None,
    include_private_values=False,
    write_workbook=True,
    write_csvs=True,
    allow_custom_output_dir=False,
):
    output_root = _normalize_output_dir(output_dir, allow_custom_output_dir)
    rows_by_sheet = build_ratecon_review_v2_rows(
        measurement_rows,
        local_document_names_by_alias=local_document_names_by_alias,
        include_private_values=include_private_values,
    )
    paths = {}
    csv_specs = {
        "instructions_csv": (
            review_v2_instructions_csv_path,
            SHEET_V2_INSTRUCTIONS,
            V2_INSTRUCTIONS_COLUMNS,
        ),
        "document_summary_csv": (
            review_v2_document_summary_csv_path,
            SHEET_V2_DOCUMENT_SUMMARY,
            V2_DOCUMENT_SUMMARY_COLUMNS,
        ),
        "core_fields_csv": (
            review_v2_core_fields_csv_path,
            SHEET_V2_CORE_FIELDS,
            V2_CORE_FIELD_COLUMNS,
        ),
        "stops_csv": (
            review_v2_stops_csv_path,
            SHEET_V2_STOPS,
            V2_STOP_COLUMNS,
        ),
        "rates_csv": (
            review_v2_rates_csv_path,
            SHEET_V2_RATES,
            V2_RATE_COLUMNS,
        ),
        "load_ids_csv": (
            review_v2_load_ids_csv_path,
            SHEET_V2_LOAD_IDS,
            V2_LOAD_ID_COLUMNS,
        ),
    }
    if write_csvs:
        for key, (path_builder, sheet_name, columns) in csv_specs.items():
            path = path_builder(output_root)
            _write_csv(path, rows_by_sheet.get(sheet_name, []), columns)
            paths[key] = path
    xlsx_written = False
    if write_workbook:
        xlsx_path = review_v2_workbook_path(output_root)
        xlsx_written = _write_v2_xlsx_if_available(xlsx_path, rows_by_sheet)
        if xlsx_written:
            paths["review_v2_workbook_xlsx"] = xlsx_path
    return {
        "paths": paths,
        "rows_by_sheet": rows_by_sheet,
        "summary": summarize_ratecon_review_v2_rows(rows_by_sheet),
        "xlsx_written": xlsx_written,
        "csvs_written": bool(write_csvs),
        "include_private_values_local_only": bool(include_private_values),
        "local_only": True,
        "raw_text_included": False,
        "private_values_printed": False,
    }


def write_ratecon_review_v2_rows_artifacts(
    rows_by_sheet,
    output_dir=None,
    write_workbook=True,
    write_csvs=True,
    allow_custom_output_dir=False,
):
    output_root = _normalize_output_dir(output_dir, allow_custom_output_dir)
    rows_by_sheet = rows_by_sheet or {}
    paths = {}
    csv_specs = {
        "instructions_csv": (
            review_v2_instructions_csv_path,
            SHEET_V2_INSTRUCTIONS,
            V2_INSTRUCTIONS_COLUMNS,
        ),
        "document_summary_csv": (
            review_v2_document_summary_csv_path,
            SHEET_V2_DOCUMENT_SUMMARY,
            V2_DOCUMENT_SUMMARY_COLUMNS,
        ),
        "core_fields_csv": (
            review_v2_core_fields_csv_path,
            SHEET_V2_CORE_FIELDS,
            V2_CORE_FIELD_COLUMNS,
        ),
        "stops_csv": (
            review_v2_stops_csv_path,
            SHEET_V2_STOPS,
            V2_STOP_COLUMNS,
        ),
        "rates_csv": (
            review_v2_rates_csv_path,
            SHEET_V2_RATES,
            V2_RATE_COLUMNS,
        ),
        "load_ids_csv": (
            review_v2_load_ids_csv_path,
            SHEET_V2_LOAD_IDS,
            V2_LOAD_ID_COLUMNS,
        ),
    }
    if write_csvs:
        for key, (path_builder, sheet_name, columns) in csv_specs.items():
            path = path_builder(output_root)
            _write_csv(path, rows_by_sheet.get(sheet_name, []), columns)
            paths[key] = path
    xlsx_written = False
    if write_workbook:
        xlsx_path = review_v2_workbook_path(output_root)
        xlsx_written = _write_v2_xlsx_if_available(xlsx_path, rows_by_sheet)
        if xlsx_written:
            paths["review_v2_workbook_xlsx"] = xlsx_path
    return {
        "paths": paths,
        "rows_by_sheet": rows_by_sheet,
        "summary": summarize_ratecon_review_v2_rows(rows_by_sheet),
        "xlsx_written": xlsx_written,
        "csvs_written": bool(write_csvs),
        "local_only": True,
        "raw_text_included": False,
        "private_values_printed": False,
    }


def write_ratecon_review_artifacts(
    measurement_rows,
    output_dir=None,
    local_document_names_by_alias=None,
    include_private_values=False,
    write_workbook=True,
    write_csvs=True,
    allow_custom_output_dir=False,
):
    rows_by_sheet = build_ratecon_review_rows(
        measurement_rows,
        local_document_names_by_alias=local_document_names_by_alias,
        include_private_values=include_private_values,
    )
    output_root = _normalize_output_dir(output_dir, allow_custom_output_dir)
    paths = {}
    csv_result = None
    if write_csvs:
        csv_result = write_ratecon_review_csvs(
            measurement_rows,
            output_dir=output_root,
            local_document_names_by_alias=local_document_names_by_alias,
            include_private_values=include_private_values,
            allow_custom_output_dir=True,
        )
        paths.update(csv_result["paths"])
    workbook_result = None
    if write_workbook:
        workbook_result = write_ratecon_review_workbook(
            measurement_rows,
            output_dir=output_root,
            local_document_names_by_alias=local_document_names_by_alias,
            include_private_values=include_private_values,
            allow_custom_output_dir=True,
        )
        if workbook_result.get("xlsx"):
            paths["review_workbook_xlsx"] = workbook_result["xlsx"]

    return {
        "paths": paths,
        "rows_by_sheet": rows_by_sheet,
        "summary": summarize_ratecon_review_workbook_rows(rows_by_sheet),
        "xlsx_written": bool(workbook_result and workbook_result.get("xlsx")),
        "csvs_written": bool(csv_result),
        "include_private_values_local_only": bool(include_private_values),
        "local_only": True,
        "raw_text_included": False,
        "private_values_printed": False,
    }
