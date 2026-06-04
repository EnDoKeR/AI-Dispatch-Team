"""Local-only Google Sheets compatible review export for private measurement."""

import csv
from pathlib import Path

from app.document_ai.private_measurement_outputs import (
    _candidate_counts_summary,
    _join,
    _normalize_output_dir,
)
from app.document_ai.measurement_cli.ratecon_private_output_paths import (
    REVIEW_GOOGLE_SHEET_CSV,
    REVIEW_WORKBOOK_XLSX,
    review_google_sheet_csv_path,
    review_workbook_path,
)


REVIEW_EXPORT_COLUMNS = [
    "Folder Order",
    "Local Document Name / File Stem",
    "Measurement Alias",
    "Document Type",
    "Classification Status",
    "Extraction Relevant",
    "Normal Load Movement",
    "OCR Needed",
    "Layout Attempted",
    "Provider Status",
    "Raw Stop Groups",
    "Premerge Groups",
    "Post Single-Line Cluster",
    "Post Row Merge",
    "Post Section Cluster",
    "Post Noise Filter",
    "Post Dedupe",
    "Normalized Stops",
    "Old Raw Stop Groups",
    "Old Normalized Stops",
    "Stop Span Anchors",
    "Stop Spans",
    "Span Normalized Stops",
    "Span Pickup Count",
    "Span Delivery Count",
    "Span Generic Stop Count",
    "Span Unknown Count",
    "Span Date Resolved",
    "Span Date Missing",
    "Span Time Resolved",
    "Span Time Missing",
    "Span Review Required Count",
    "Stop Span Passthrough Detected",
    "Old vs Span Delta",
    "Recommended Review Priority",
    "Duplicate Removed",
    "Noise Removed",
    "Pickup Count",
    "Delivery Count",
    "Generic Stop Count",
    "Unknown Stop Count",
    "Date Resolved",
    "Date Missing",
    "Time Resolved",
    "Time Missing",
    "Rate Status",
    "Broker Identity Status",
    "Equipment Status",
    "Weight Status",
    "Top Blocker",
    "Root Cause Bucket",
    "Recommended Engineering Action",
    "Review Status",
    "User Notes Local Only",
]


def _text(value):
    return str(value or "").strip()


def _field_status(row, field_name):
    for field_status in row.get("field_statuses", []) or []:
        if not isinstance(field_status, dict):
            continue
        if field_status.get("field_name") == field_name:
            return _text(field_status.get("status"))
    return ""


def _broker_identity_status(row):
    statuses = [
        _field_status(row, "broker_name"),
        _field_status(row, "broker_mc"),
    ]
    statuses = [status for status in statuses if status]
    if not statuses:
        return ""
    if all(status == "resolved" for status in statuses):
        return "resolved"
    if any(status == "conflict" for status in statuses):
        return "conflict"
    if any(status == "missing" for status in statuses):
        return "missing"
    return ";".join(sorted(set(statuses)))


def _count_stop_field_status(row, field_name, status):
    counts = row.get("stop_field_status_counts", {})
    if not isinstance(counts, dict):
        return 0
    field_counts = counts.get(field_name, {})
    if not isinstance(field_counts, dict):
        return 0
    return int(field_counts.get(status, 0) or 0)


def _root_cause_bucket(row):
    trace = row.get("stop_pipeline_trace", {}) or {}
    if trace.get("passthrough_detected"):
        return "NORMALIZER_PASSTHROUGH"
    first_changed = _text(trace.get("first_stage_that_changed"))
    if first_changed:
        return first_changed
    if int(row.get("raw_stop_group_count", 0) or 0) == int(row.get("normalized_stop_count", 0) or 0):
        return "NORMALIZER_PASSTHROUGH"
    return _text(row.get("layout_likely_issue_bucket")) or "review_needed"


def _recommended_action(row):
    root = _root_cause_bucket(row)
    if root == "NORMALIZER_PASSTHROUGH":
        return "inspect stop grouping wiring"
    if root == "post_single_line_cluster":
        return "review clustered stop values locally"
    if int(row.get("post_single_line_cluster_stop_group_count", 0) or 0) < int(row.get("premerge_stop_group_count", 0) or 0):
        return "review normalized stop correctness"
    if int(row.get("date_candidate_attached_count", 0) or 0) == 0:
        return "inspect date/time attachment"
    return "review blocker bucket"


def _old_vs_span_delta(row):
    old_count = int(row.get("old_normalized_stops", row.get("normalized_stop_count", 0)) or 0)
    span_count = int(row.get("span_normalized_stop_count", 0) or 0)
    return old_count - span_count


def _recommended_review_priority(row):
    if row.get("span_passthrough_detected"):
        return "high"
    if int(row.get("span_normalized_stop_count", 0) or 0) == 0 and row.get("layout_provider_status") == "success":
        return "high"
    if int(row.get("span_review_required_count", 0) or 0):
        return "medium"
    if _old_vs_span_delta(row) > 0:
        return "value_review"
    return "normal"


def build_review_export_rows(rows, local_document_names_by_alias=None):
    local_names = local_document_names_by_alias or {}
    export_rows = []
    for index, row in enumerate(rows or [], start=1):
        alias = _text(row.get("document_alias"))
        blockers = row.get("blocker_categories", []) or []
        export_rows.append(
            {
                "Folder Order": index,
                "Local Document Name / File Stem": _text(local_names.get(alias)),
                "Measurement Alias": alias,
                "Document Type": _text(row.get("document_type")),
                "Classification Status": _text(row.get("classification_status")),
                "Extraction Relevant": bool(row.get("extraction_relevant")),
                "Normal Load Movement": bool(row.get("normal_load_movement")),
                "OCR Needed": row.get("extraction_status") == "EMPTY_TEXT",
                "Layout Attempted": bool(_text(row.get("layout_provider_status")))
                and not _text(row.get("layout_provider_status")).startswith("skipped"),
                "Provider Status": _text(row.get("layout_provider_status")),
                "Raw Stop Groups": int(row.get("raw_stop_group_count", 0) or 0),
                "Premerge Groups": int(row.get("premerge_stop_group_count", 0) or 0),
                "Post Single-Line Cluster": int(row.get("post_single_line_cluster_stop_group_count", 0) or 0),
                "Post Row Merge": int(row.get("post_row_merge_stop_group_count", 0) or 0),
                "Post Section Cluster": int(row.get("post_section_merge_stop_group_count", 0) or 0),
                "Post Noise Filter": int(row.get("post_noise_filter_stop_group_count", 0) or 0),
                "Post Dedupe": int(row.get("post_dedupe_stop_group_count", 0) or 0),
                "Normalized Stops": int(row.get("normalized_stop_count", 0) or 0),
                "Old Raw Stop Groups": int(row.get("old_raw_stop_groups", row.get("raw_stop_group_count", 0)) or 0),
                "Old Normalized Stops": int(row.get("old_normalized_stops", row.get("normalized_stop_count", 0)) or 0),
                "Stop Span Anchors": int(row.get("span_anchor_count", 0) or 0),
                "Stop Spans": int(row.get("stop_span_count", 0) or 0),
                "Span Normalized Stops": int(row.get("span_normalized_stop_count", 0) or 0),
                "Span Pickup Count": int(row.get("span_pickup_count", 0) or 0),
                "Span Delivery Count": int(row.get("span_delivery_count", 0) or 0),
                "Span Generic Stop Count": int(row.get("span_generic_stop_count", 0) or 0),
                "Span Unknown Count": int(row.get("span_unknown_count", 0) or 0),
                "Span Date Resolved": int(row.get("span_date_resolved_count", 0) or 0),
                "Span Date Missing": int(row.get("span_date_missing_count", 0) or 0),
                "Span Time Resolved": int(row.get("span_time_resolved_count", 0) or 0),
                "Span Time Missing": int(row.get("span_time_missing_count", 0) or 0),
                "Span Review Required Count": int(row.get("span_review_required_count", 0) or 0),
                "Stop Span Passthrough Detected": bool(row.get("span_passthrough_detected")),
                "Old vs Span Delta": _old_vs_span_delta(row),
                "Recommended Review Priority": _recommended_review_priority(row),
                "Duplicate Removed": int(row.get("stop_duplicate_removed_count", 0) or 0),
                "Noise Removed": int(row.get("stop_noise_removed_count", 0) or 0),
                "Pickup Count": int(row.get("pickup_count", 0) or 0),
                "Delivery Count": int(row.get("delivery_count", 0) or 0),
                "Generic Stop Count": int(row.get("generic_stop_count", 0) or 0),
                "Unknown Stop Count": int(row.get("unknown_stop_count", 0) or 0),
                "Date Resolved": _count_stop_field_status(row, "date", "resolved"),
                "Date Missing": _count_stop_field_status(row, "date", "missing"),
                "Time Resolved": _count_stop_field_status(row, "time", "resolved"),
                "Time Missing": _count_stop_field_status(row, "time", "missing"),
                "Rate Status": _field_status(row, "rate"),
                "Broker Identity Status": _broker_identity_status(row),
                "Equipment Status": _field_status(row, "equipment"),
                "Weight Status": _field_status(row, "weight"),
                "Top Blocker": _text(blockers[0] if blockers else ""),
                "Root Cause Bucket": _root_cause_bucket(row),
                "Recommended Engineering Action": _recommended_action(row),
                "Review Status": "",
                "User Notes Local Only": "",
            }
        )
    return export_rows


def _write_csv(path, rows):
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=REVIEW_EXPORT_COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def _write_xlsx_if_available(path, rows):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception:
        return False

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "RateCon Review"
    sheet.append(REVIEW_EXPORT_COLUMNS)
    for row in rows:
        sheet.append([row.get(column, "") for column in REVIEW_EXPORT_COLUMNS])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(REVIEW_EXPORT_COLUMNS, start=1):
        width = min(max(len(column) + 2, 12), 36)
        sheet.column_dimensions[get_column_letter(index)].width = width

    workbook.save(path)
    return True


def write_ratecon_review_export(
    rows,
    output_dir=None,
    local_document_names_by_alias=None,
    allow_custom_output_dir=False,
):
    output_root = _normalize_output_dir(output_dir, allow_custom_output_dir)
    export_rows = build_review_export_rows(
        rows,
        local_document_names_by_alias=local_document_names_by_alias,
    )
    csv_path = review_google_sheet_csv_path(output_root)
    xlsx_path = review_workbook_path(output_root)
    _write_csv(csv_path, export_rows)
    xlsx_written = _write_xlsx_if_available(xlsx_path, export_rows)

    return {
        "csv": csv_path,
        "xlsx": xlsx_path if xlsx_written else None,
        "row_count": len(export_rows),
        "raw_text_saved": False,
        "private_values_redacted": True,
        "local_only": True,
    }
